"""
SnackZone - Office Cafeteria Ordering App
Flask backend with Excel storage + UPI Payment
"""
from flask import Flask, request, redirect, url_for, session, jsonify, send_from_directory
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime
import os
import json
import base64

app = Flask(__name__, template_folder='.', static_folder='.', static_url_path='')
app.secret_key = 'snackzone-secret-key-change-in-production'

# ─── Files ────────────────────────────────────────────────────────
USERS_FILE    = 'users.xlsx'
ORDERS_FILE   = 'orders.xlsx'
PRODUCTS_FILE = 'products.xlsx'
PAYMENT_FILE  = 'payment_settings.json'

# ─── Default Admin ────────────────────────────────────────────────
ADMIN_EMAIL    = 'admin@snackzone.com'
ADMIN_PASSWORD = 'admin123'

# ================================================================
#  Payment Settings (UPI)
# ================================================================

def get_payment_settings():
    if not os.path.exists(PAYMENT_FILE):
        return {'upi_id': '', 'upi_name': '', 'qr_image': ''}
    with open(PAYMENT_FILE, 'r') as f:
        return json.load(f)

def save_payment_settings(data):
    with open(PAYMENT_FILE, 'w') as f:
        json.dump(data, f)

# ================================================================
#  Excel helpers — Users
# ================================================================

def init_users_excel():
    if not os.path.exists(USERS_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Users'
        ws.append(['ID', 'Name', 'Email', 'Phone', 'Password_Hash', 'Role', 'Created_At'])
        for cell in ws[1]: cell.font = Font(bold=True)
        # Create default admin
        ws.append([1, 'Admin', ADMIN_EMAIL, '', generate_password_hash(ADMIN_PASSWORD), 'admin',
                   datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        wb.save(USERS_FILE)
        print(f'✓ Created {USERS_FILE} with default admin')


def get_all_users():
    if not os.path.exists(USERS_FILE):
        init_users_excel()
    wb = load_workbook(USERS_FILE)
    ws = wb.active

    # Detect schema: check if column 6 header is 'Role'
    header = [cell.value for cell in ws[1]]
    has_role_col = len(header) >= 7 and str(header[5] or '').lower() == 'role'

    users = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        if has_role_col:
            role       = row[5] if row[5] else 'user'
            created_at = row[6]
        else:
            # Old 6-column schema: no Role column
            role       = 'user'
            created_at = row[5]

        users.append({
            'id': row[0], 'name': row[1], 'email': row[2],
            'phone': row[3], 'password_hash': row[4],
            'role': role,
            'created_at': created_at
        })
    return users


def find_user_by_email(email):
    email = email.lower().strip()
    for u in get_all_users():
        if u['email'] and str(u['email']).lower() == email:
            return u
    return None


def add_user(name, email, phone, password, role='user'):
    init_users_excel()
    wb = load_workbook(USERS_FILE)
    ws = wb.active
    # Find the highest existing numeric ID
    max_id = 1
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0] is not None:
            try:
                max_id = max(max_id, int(row[0]))
            except (ValueError, TypeError):
                pass
    new_id = max_id + 1
    ws.append([
        new_id, name.strip(), email.lower().strip(),
        phone.strip() if phone else '',
        generate_password_hash(password),
        role,
        datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ])
    wb.save(USERS_FILE)
    return new_id

# ================================================================
#  Excel helpers — Products
# ================================================================

def init_products_excel():
    if not os.path.exists(PRODUCTS_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Products'
        ws.append(['ID', 'Name', 'Category', 'Price', 'Description', 'Emoji', 'Available', 'Created_At'])
        for cell in ws[1]: cell.font = Font(bold=True)
        # Default snacks
        defaults = [
            [1, 'Samosa', 'Snacks', 10, 'Crispy fried samosa', '🥟', True],
            [2, 'Vada Pav', 'Snacks', 15, 'Mumbai style vada pav', '🍔', True],
            [3, 'Tea', 'Beverages', 10, 'Hot masala chai', '☕', True],
            [4, 'Coffee', 'Beverages', 20, 'Filter coffee', '☕', True],
            [5, 'Biscuits', 'Snacks', 10, 'Assorted biscuit pack', '🍪', True],
            [6, 'Cold Drink', 'Beverages', 25, '250ml cold drink bottle', '🥤', True],
        ]
        for d in defaults:
            ws.append(d + [datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        wb.save(PRODUCTS_FILE)
        print(f'✓ Created {PRODUCTS_FILE}')


def get_all_products():
    if not os.path.exists(PRODUCTS_FILE):
        init_products_excel()
    wb = load_workbook(PRODUCTS_FILE)
    ws = wb.active
    products = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0] is not None:
            products.append({
                'id': row[0], 'name': row[1], 'category': row[2],
                'price': row[3], 'description': row[4],
                'emoji': row[5] if row[5] else '🍽️',
                'available': row[6] if row[6] is not None else True,
            })
    return products


def add_product(name, category, price, description, emoji):
    init_products_excel()
    wb = load_workbook(PRODUCTS_FILE)
    ws = wb.active
    new_id = ws.max_row  # simple id
    ws.append([
        new_id, name.strip(), category.strip(),
        float(price), description.strip(), emoji.strip(), True,
        datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ])
    wb.save(PRODUCTS_FILE)
    return new_id


def update_product_availability(product_id, available):
    wb = load_workbook(PRODUCTS_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if row[0].value == int(product_id):
            row[6].value = available
            break
    wb.save(PRODUCTS_FILE)


def delete_product(product_id):
    wb = load_workbook(PRODUCTS_FILE)
    ws = wb.active
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if row[0].value == int(product_id):
            ws.delete_rows(i)
            break
    wb.save(PRODUCTS_FILE)

# ================================================================
#  Excel helpers — Orders
# ================================================================

def init_orders_excel():
    if not os.path.exists(ORDERS_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Orders'
        ws.append(['Order_ID', 'User_Name', 'User_Email', 'Items', 'Total_Amount', 'Status', 'UTR_Number', 'Created_At'])
        for cell in ws[1]: cell.font = Font(bold=True)
        wb.save(ORDERS_FILE)
        print(f'✓ Created {ORDERS_FILE}')


def save_order(user, items, amount, utr='', status='PENDING'):
    init_orders_excel()
    wb = load_workbook(ORDERS_FILE)
    ws = wb.active
    # Find max existing ID
    max_id = 1
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0] is not None:
            try: max_id = max(max_id, int(row[0]))
            except: pass
    order_id = max_id + 1
    ws.append([
        order_id,
        user.get('name', ''),
        user.get('email', ''),
        json.dumps(items, ensure_ascii=False),   # always store as JSON string
        float(amount),
        status,
        str(utr),
        datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ])
    wb.save(ORDERS_FILE)
    return order_id


def get_all_orders():
    if not os.path.exists(ORDERS_FILE):
        init_orders_excel()
        return []
    wb = load_workbook(ORDERS_FILE)
    ws = wb.active

    # Detect schema by header
    header = [str(c.value or '').strip() for c in ws[1]]
    # New schema: Order_ID, User_Name, User_Email, Items, Total_Amount, Status, UTR_Number, Created_At
    # Old Razorpay schema: Order_ID, Razorpay_Order_ID, Payment_ID, User_Name, User_Email, Items, Amount_INR, Status, Created_At
    is_old = len(header) >= 9 and 'razorpay' in header[1].lower()

    orders = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        try:
            if is_old:
                items = json.loads(row[5]) if row[5] else []
                entry = {
                    'order_id':   row[0],
                    'user_name':  row[3],
                    'user_email': row[4],
                    'items':      items,
                    'amount':     row[6],
                    'status':     row[7] or 'PENDING',
                    'utr':        row[2] or '',
                    'created_at': str(row[8] or ''),
                }
            else:
                items = json.loads(row[3]) if row[3] else []
                entry = {
                    'order_id':   row[0],
                    'user_name':  row[1],
                    'user_email': row[2],
                    'items':      items,
                    'amount':     row[4],
                    'status':     row[5] or 'PENDING',
                    'utr':        row[6] or '',
                    'created_at': str(row[7] or ''),
                }
            orders.append(entry)
        except Exception as e:
            print(f'Warning: skipping malformed order row: {e}')
            continue

    return list(reversed(orders))  # newest first


def update_order_status(order_id, status):
    wb = load_workbook(ORDERS_FILE)
    ws = wb.active
    # Find the status column from header
    header = [str(c.value or '').lower() for c in ws[1]]
    status_col = None
    for i, h in enumerate(header):
        if h == 'status':
            status_col = i
            break
    if status_col is None:
        status_col = 5  # fallback
    for row in ws.iter_rows(min_row=2):
        if row[0].value == int(order_id):
            row[status_col].value = status
            break
    wb.save(ORDERS_FILE)

# ================================================================
#  Page Routes
# ================================================================

@app.route('/')
def home():
    if 'user' not in session:
        return redirect('/login')
    if session['user'].get('role') == 'admin':
        return redirect('/admin')
    return send_from_directory('.', 'index.html')

@app.route('/login')
def login_page():
    return send_from_directory('.', 'login.html')

@app.route('/admin')
def admin_page():
    if 'user' not in session or session['user'].get('role') != 'admin':
        return redirect('/login')
    return send_from_directory('.', 'admin.html')

@app.route('/checkout.html')
def checkout_page():
    if 'user' not in session:
        return redirect('/login')
    return send_from_directory('.', 'checkout.html')

@app.route('/index.html')
def index_html():
    return redirect('/')

# ================================================================
#  Auth API
# ================================================================

@app.route('/api/signup', methods=['POST'])
def api_signup():
    data = request.get_json() or {}
    name     = (data.get('name') or '').strip()
    email    = (data.get('email') or '').strip()
    phone    = (data.get('phone') or '').strip()
    password = data.get('password') or ''

    if not name or not email or not password:
        return jsonify({'success': False, 'message': 'Name, email and password are required'}), 400
    if len(password) < 6:
        return jsonify({'success': False, 'message': 'Password must be at least 6 characters'}), 400
    if find_user_by_email(email):
        return jsonify({'success': False, 'message': 'This email is already registered'}), 400

    uid = add_user(name, email, phone, password, role='user')
    session['user'] = {'id': uid, 'name': name, 'email': email.lower(), 'role': 'user'}
    return jsonify({'success': True, 'message': f'Welcome {name}!', 'redirect': '/'})


@app.route('/api/login', methods=['POST'])
def api_login():
    data     = request.get_json() or {}
    email    = (data.get('email') or '').strip()
    password = data.get('password') or ''

    if not email or not password:
        return jsonify({'success': False, 'message': 'Email and password required'}), 400

    user = find_user_by_email(email)
    if not user or not check_password_hash(user['password_hash'], password):
        return jsonify({'success': False, 'message': 'Incorrect email or password'}), 401

    session['user'] = {
        'id': user['id'], 'name': user['name'],
        'email': user['email'], 'role': user.get('role', 'user')
    }
    redirect_url = '/admin' if user.get('role') == 'admin' else '/'
    return jsonify({'success': True, 'message': f'Welcome, {user["name"]}!', 'redirect': redirect_url})


@app.route('/api/logout', methods=['POST'])
def api_logout():
    session.pop('user', None)
    return jsonify({'success': True, 'redirect': '/login'})


@app.route('/api/me')
def api_me():
    if 'user' in session:
        return jsonify({'logged_in': True, 'user': session['user']})
    return jsonify({'logged_in': False})

# ================================================================
#  Products API
# ================================================================

@app.route('/api/products', methods=['GET'])
def api_get_products():
    products = get_all_products()
    return jsonify({'success': True, 'products': products})


@app.route('/api/admin/products', methods=['POST'])
def api_add_product():
    if 'user' not in session or session['user'].get('role') != 'admin':
        return jsonify({'success': False, 'message': 'Admin access required'}), 403
    data = request.get_json() or {}
    name        = (data.get('name') or '').strip()
    category    = (data.get('category') or 'Snacks').strip()
    price       = data.get('price', 0)
    description = (data.get('description') or '').strip()
    emoji       = (data.get('emoji') or '🍽️').strip()

    if not name or not price:
        return jsonify({'success': False, 'message': 'Name and price are required'}), 400

    pid = add_product(name, category, price, description, emoji)
    return jsonify({'success': True, 'message': f'Product "{name}" added!', 'id': pid})


@app.route('/api/admin/products/<int:product_id>', methods=['DELETE'])
def api_delete_product(product_id):
    if 'user' not in session or session['user'].get('role') != 'admin':
        return jsonify({'success': False, 'message': 'Admin access required'}), 403
    delete_product(product_id)
    return jsonify({'success': True, 'message': 'Product deleted'})


@app.route('/api/admin/products/<int:product_id>/toggle', methods=['POST'])
def api_toggle_product(product_id):
    if 'user' not in session or session['user'].get('role') != 'admin':
        return jsonify({'success': False, 'message': 'Admin access required'}), 403
    data = request.get_json() or {}
    available = data.get('available', True)
    update_product_availability(product_id, available)
    return jsonify({'success': True})

# ================================================================
#  Orders API
# ================================================================

@app.route('/api/orders', methods=['POST'])
def api_place_order():
    if 'user' not in session:
        return jsonify({'success': False, 'message': 'Login required'}), 401
    data   = request.get_json() or {}
    items  = data.get('items', [])
    amount = data.get('amount', 0)
    utr    = (data.get('utr') or '').strip()

    if not items or amount <= 0:
        return jsonify({'success': False, 'message': 'Cart is empty'}), 400

    order_id = save_order(session['user'], items, amount, utr=utr, status='PENDING')
    return jsonify({'success': True, 'message': 'Order placed successfully!', 'order_id': order_id})


@app.route('/api/admin/add-order', methods=['POST'])
def api_admin_add_order():
    """Admin places a manual order on behalf of a customer."""
    if 'user' not in session or session['user'].get('role') != 'admin':
        return jsonify({'success': False, 'message': 'Admin access required'}), 403
    data  = request.get_json() or {}
    name  = (data.get('customer_name') or '').strip()
    email = (data.get('customer_email') or '').strip()
    items  = data.get('items', [])
    amount = data.get('amount', 0)
    status = data.get('status', 'PENDING')
    utr    = (data.get('utr') or '').strip()

    if not name or not items or amount <= 0:
        return jsonify({'success': False, 'message': 'Customer name and items are required'}), 400

    user_obj = {'name': name, 'email': email or f'{name.lower().replace(" ",".")}@office'}
    order_id = save_order(user_obj, items, amount, utr=utr, status=status)
    return jsonify({'success': True, 'message': 'Order placed!', 'order_id': order_id})


@app.route('/api/my-orders', methods=['GET'])
def api_my_orders():
    """Return orders for the currently logged-in user."""
    if 'user' not in session:
        return jsonify({'success': False, 'message': 'Login required'}), 401
    email = session['user'].get('email', '').lower()
    all_orders = get_all_orders()
    my = [o for o in all_orders if str(o.get('user_email', '')).lower() == email]
    return jsonify({'success': True, 'orders': my})


@app.route('/api/admin/orders', methods=['GET'])
def api_get_orders():
    if 'user' not in session or session['user'].get('role') != 'admin':
        return jsonify({'success': False, 'message': 'Admin access required'}), 403
    orders = get_all_orders()
    # Guarantee items is always a list, never a raw string
    for o in orders:
        if isinstance(o.get('items'), str):
            try:    o['items'] = json.loads(o['items'])
            except: o['items'] = []
        elif not isinstance(o.get('items'), list):
            o['items'] = []
    return jsonify({'success': True, 'orders': orders})


@app.route('/api/admin/orders/<int:order_id>/status', methods=['POST'])
def api_update_order_status(order_id):
    if 'user' not in session or session['user'].get('role') != 'admin':
        return jsonify({'success': False, 'message': 'Admin access required'}), 403
    data   = request.get_json() or {}
    status = data.get('status', 'PENDING')
    update_order_status(order_id, status)
    return jsonify({'success': True})

# ================================================================
#  Payment Settings API
# ================================================================

@app.route('/api/payment-settings', methods=['GET'])
def api_get_payment():
    settings = get_payment_settings()
    # Don't expose QR image data to non-logged-in users on GET
    if 'user' not in session:
        return jsonify({'success': False, 'message': 'Login required'}), 401
    return jsonify({'success': True, 'settings': settings})


@app.route('/api/admin/payment-settings', methods=['POST'])
def api_save_payment():
    if 'user' not in session or session['user'].get('role') != 'admin':
        return jsonify({'success': False, 'message': 'Admin access required'}), 403
    data = request.get_json() or {}
    current = get_payment_settings()
    current['upi_id']   = (data.get('upi_id') or '').strip()
    current['upi_name'] = (data.get('upi_name') or '').strip()
    if data.get('qr_image'):
        current['qr_image'] = data['qr_image']
    save_payment_settings(current)
    return jsonify({'success': True, 'message': 'Payment settings saved!'})

# ================================================================
#  Admin stats
# ================================================================

@app.route('/api/admin/stats', methods=['GET'])
def api_admin_stats():
    if 'user' not in session or session['user'].get('role') != 'admin':
        return jsonify({'success': False}), 403
    orders   = get_all_orders()
    products = get_all_products()
    users    = get_all_users()
    total_revenue = sum(o['amount'] for o in orders if o['status'] in ('PAID', 'CONFIRMED'))
    pending_count = sum(1 for o in orders if o['status'] == 'PENDING')
    return jsonify({
        'success': True,
        'stats': {
            'total_orders': len(orders),
            'pending_orders': pending_count,
            'total_products': len(products),
            'total_users': len(users),
            'total_revenue': total_revenue,
        }
    })

# ================================================================
#  Run
# ================================================================
if __name__ == '__main__':
    init_users_excel()
    init_products_excel()
    init_orders_excel()
    print('=' * 55)
    print('  🍔 SnackZone - Office Cafeteria')
    print('  Login page:  http://localhost:5000/login')
    print('  User home:   http://localhost:5000/')
    print('  Admin panel: http://localhost:5000/admin')
    print(f'  Admin login: {ADMIN_EMAIL} / {ADMIN_PASSWORD}')
    print('=' * 55)
    app.run(debug=True, port=5000)
